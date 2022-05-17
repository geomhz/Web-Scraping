from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.common.exceptions import *
from time import sleep
import openpyxl
from loguru import logger as lg

class Brisbane_Apto:

    def __init__(self):
        lg.success("Iniciando pesquisa de preços")
        self.proxima_pagina = 1
        self.nome_ruas = []
        self.valores = []
        self.quarto = []
        self.ref_links = []
        sleep(2)
        options = Options()
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        options.add_experimental_option('prefs', {
            "safebrowsing.enabled": True
        })
        options.add_argument("--start-maximized")
        self.driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)

    def iniciar(self):
        self.carrega_pagina_web()
        self.informacoes_casas()
        self.planilha()

    def carrega_pagina_web(self) -> None:
        self.wait = WebDriverWait(self.driver, 2)
        self.wait2 = WebDriverWait(self.driver, 120)
        self.driver.get(f"https://www.rent.com.au/properties/brisbane-qld-4000/p{self.proxima_pagina}")
        sleep(2)

    def informacoes_casas(self):

        ruas = self.driver.find_elements_by_xpath('//h2[@class="address"]')
        precos = self.driver.find_elements_by_xpath('//span[@class="price"]')
        camas = self.driver.find_elements_by_xpath('//span[@class="value tx-normal-lead -block"]')
        links = self.driver.find_elements_by_xpath('//a[@class="link"]')
        

        self.nome_ruas = self.listar_informacoes(self.nome_ruas, ruas, "texto")
        self.valores = self.listar_informacoes(self.valores, precos, "texto")
        self.quarto = self.listar_informacoes(self.quarto, camas, "texto")
        self.ref_links = self.listar_informacoes(self.ref_links, links, "atributo")
        self.prox_pagina(ruas)

    def listar_informacoes(self, lista, prop, tipo):
        if tipo == "texto":
            for i in range(len(prop)):
                lista.append(prop[i].text)
        elif tipo == "atributo":
            for i in range(len(prop)):
                lista.append(prop[i].get_attribute("href"))                            
        return lista 

    def prox_pagina(self, ruas):
        self.proxima_pagina += 1
        if len(ruas) == 0 : 
            lg.success("Fim do escaneamento.")
            lg.success("Não há mais apartamentos para alugar. Fim do escaneamento.")
            sleep(1)
            lg.debug("Fechando navegador em 3s.")
            sleep(1)
            lg.debug("Fechando navegador em 2s.")
            sleep(1)
            lg.debug("Fechando navegador em 1s.")
            sleep(1)
            self.driver.quit()
        else:
            pag = (f"https://www.rent.com.au/properties/brisbane-qld-4000/p{self.proxima_pagina}")
            lg.info("Acessando página {}".format(self.proxima_pagina))
            self.driver.get(pag)
            self.informacoes_casas()
            
    def planilha(self):
            index = 2
            self.planilha = openpyxl.Workbook()
            aptos = self.planilha['Sheet']
            aptos.title = 'Brisbane'
            aptos['A1'] = 'Rua'
            aptos['B1'] = 'Preço e Tipo'
            aptos['C1'] = 'Quartos'
            aptos['D1'] = 'Link'

            for ruas, precos, camas, links in zip(self.nome_ruas, self.valores, self.quarto, self.ref_links):
                aptos.cell(column=1, row=index, value=ruas)
                aptos.cell(column=2, row=index, value=precos)
                aptos.cell(column=3, row=index, value=camas)
                aptos.cell(column=4, row=index, value=links)
                index += 1

            self.planilha.save("aptos.xlsx")
            lg.success("Relatório criado com sucesso")


if __name__ == '__main__':
    go = Brisbane_Apto()
    go.iniciar()